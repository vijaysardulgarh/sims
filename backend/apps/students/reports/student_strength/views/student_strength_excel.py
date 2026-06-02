from rest_framework.views import APIView

from django.http import HttpResponse

import openpyxl

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
)

from .student_strength import (
    StudentStrengthAPIView
)


class StudentStrengthExcelAPIView(
    APIView
):

    def get(self, request):

        workbook = openpyxl.Workbook()

        sheet = workbook.active

        sheet.title = (
            "Student Strength"
        )

        headers = [

            "Class",
            "Section",

            "SC Male",
            "SC Female",
            "SC Total",

            "BC-A Male",
            "BC-A Female",
            "BC-A Total",

            "BC-B Male",
            "BC-B Female",
            "BC-B Total",

            "GEN Male",
            "GEN Female",
            "GEN Total",

            "Overall Male",
            "Overall Female",
            "Overall Total",

        ]

        sheet.append(
            headers
        )

        # =====================================
        # HEADER STYLING
        # =====================================

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78"
        )

        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        for cell in sheet[1]:

            cell.fill = (
                header_fill
            )

            cell.font = (
                header_font
            )

            cell.alignment = (
                Alignment(
                    horizontal="center"
                )
            )

        # =====================================
        # DATA
        # =====================================

        data = (
            StudentStrengthAPIView()
            .get(request)
            .data
        )

        for row in data:

            sheet.append([

                row["class_name"],

                row["section_name"],

                row["sc_male"],
                row["sc_female"],
                row["sc_total"],

                row["bca_male"],
                row["bca_female"],
                row["bca_total"],

                row["bcb_male"],
                row["bcb_female"],
                row["bcb_total"],

                row["gen_male"],
                row["gen_female"],
                row["gen_total"],

                row["overall_male"],
                row["overall_female"],
                row["overall_total"],

            ])

            current_row = (
                sheet.max_row
            )

            # ================================
            # CLASS TOTAL
            # ================================

            if (
                row["row_type"]
                == "class_total"
            ):

                for cell in sheet[
                    current_row
                ]:

                    cell.font = Font(
                        bold=True
                    )

                    cell.fill = PatternFill(
                        fill_type="solid",
                        fgColor="D9EAD3"
                    )

            # ================================
            # GRAND TOTAL
            # ================================

            elif (
                row["row_type"]
                == "grand_total"
            ):

                for cell in sheet[
                    current_row
                ]:

                    cell.font = Font(
                        bold=True
                    )

                    cell.fill = PatternFill(
                        fill_type="solid",
                        fgColor="FFF2CC"
                    )

        # =====================================
        # AUTO WIDTH
        # =====================================

        for column in sheet.columns:

            max_length = 0

            column_letter = (
                column[0].column_letter
            )

            for cell in column:

                try:

                    if (
                        len(
                            str(
                                cell.value
                            )
                        )
                        > max_length
                    ):

                        max_length = len(
                            str(
                                cell.value
                            )
                        )

                except Exception:

                    pass

            sheet.column_dimensions[
                column_letter
            ].width = (
                max_length + 3
            )

        # =====================================
        # RESPONSE
        # =====================================

        response = HttpResponse(

            content_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

        )

        response[
            "Content-Disposition"
        ] = (
            'attachment; filename="student_strength.xlsx"'
        )

        workbook.save(
            response
        )

        return response